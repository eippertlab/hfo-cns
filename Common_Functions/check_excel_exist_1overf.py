import os
import xlsxwriter
from openpyxl import load_workbook


def check_excel_exist_1overf(subjects, srmr_nr, fname, sheetname):
    if srmr_nr == 1:
        col_names = ['Subject', 'median_offset', 'median_exponent', 'tibial_offset',
                     'tibial_exponent']
    elif srmr_nr == 2:
        col_names = ['Subject', 'med_mixed_offset', 'med_mixed_exponent', 'tib_mixed_offset',
                     'tib_mixed_exponent']
    # If the excel workbook doesn't exist, create it
    if not os.path.isfile(fname):
        workbook = xlsxwriter.Workbook(fname)
        worksheet = workbook.add_worksheet(name=sheetname)
        i = 0
        for col in col_names:
            worksheet.write(0, i, col)
            i += 1
        j = 1
        for subject in subjects:
            worksheet.write(j, 0, subject)
            j += 1
        workbook.close()
    # If it does exist, check the worksheet exists, and if not, add it
    else:
        wb = load_workbook(filename=fname)
        if sheetname not in wb.sheetnames:
            worksheet = wb.create_sheet(sheetname)
            i = 1
            for col in col_names:
                worksheet.cell(row=1, column=i).value = col
                i += 1
            j = 2
            for subject in subjects:
                worksheet.cell(row=j, column=1).value = subject
                j += 1
            wb.save(fname)