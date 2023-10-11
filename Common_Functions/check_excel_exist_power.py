import os
import xlsxwriter
from openpyxl import load_workbook


def check_excel_exist_power(subjects, fname, sheetname, srmr_nr):
    if srmr_nr == 1:
        col_names = ['Subject', 'Power_median_correct', 'Power_median_incorrect', 'Power_median_ratio',
                     'Power_tibial_correct', 'Power_tibial_incorrect', 'Power_tibial_ratio']
    elif srmr_nr == 2:
        col_names = ['Subject', 'Power_med_mixed_correct', 'Power_med_mixed_incorrect', 'Power_med_mixed_ratio',
                     'Power_tib_mixed_correct', 'Power_tib_mixed_incorrect', 'Power_tib_mixed_ratio']
    # If the excel workbook doesn't exist, create it
    if not os.path.isfile(fname):
        workbook = xlsxwriter.Workbook(fname)
        worksheet = workbook.add_worksheet(name=sheetname)
        i = 0
        for col in col_names:
            worksheet.write(0, i, col)
            i += 1
        j = 1
        for subject in subjects:
            worksheet.write(j, 0, subject)
            j += 1
        workbook.close()
    # If it does exist, check the worksheet exists, and if not, add it
    else:
        wb = load_workbook(filename=fname)
        if sheetname not in wb.sheetnames:
            worksheet = wb.create_sheet(sheetname)
            i = 1
            for col in col_names:
                worksheet.cell(row=1, column=i).value = col
                i += 1
            j = 2
            for subject in subjects:
                worksheet.cell(row=j, column=1).value = subject
                j += 1
            wb.save(fname)