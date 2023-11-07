import os
import xlsxwriter
from openpyxl import load_workbook


def check_excel_exist_partialcorr(subjects, fname, sheetname):
    if sheetname == 'Cortical':
        col_names = ['Subject', 'N20', 'N20_amplitude', 'N20_high', 'N20_high_amplitude',
                     'N20_high_env', 'N20_high_amplitude_env', 'N20_SNR', 'N20_high_SNR',
                     'P39', 'P39_amplitude', 'P39_high', 'P39_high_amplitude',
                     'P39_high_env', 'P39_high_amplitude_env', 'P39_SNR', 'P39_high_SNR']
    elif sheetname == 'Spinal':
        col_names = ['Subject', 'N13', 'N13_amplitude', 'N13_high', 'N13_high_amplitude',
                     'N13_high_env', 'N13_high_amplitude_env', 'N13_SNR', 'N13_high_SNR',
                     'N22', 'N22_amplitude', 'N22_high', 'N22_high_amplitude',
                     'N22_high_env', 'N22_high_amplitude_env', 'N22_SNR', 'N22_high_SNR']
    # If the excel workbook doesn't exist, create it
    if not os.path.isfile(fname):
        workbook = xlsxwriter.Workbook(fname)
        worksheet = workbook.add_worksheet(name=sheetname)
        i = 0
        for col in col_names:
            worksheet.write(0, i, col)
            i += 1
        j = 1
        for subject in subjects:
            worksheet.write(j, 0, subject)
            j += 1
        workbook.close()
    # If it does exist, check the worksheet exists, and if not, add it
    else:
        wb = load_workbook(filename=fname)
        if sheetname not in wb.sheetnames:
            worksheet = wb.create_sheet(sheetname)
            i = 1
            for col in col_names:
                worksheet.cell(row=1, column=i).value = col
                i += 1
            j = 2
            for subject in subjects:
                worksheet.cell(row=j, column=1).value = subject
                j += 1
            wb.save(fname)