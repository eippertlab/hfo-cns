import os
import xlsxwriter
from openpyxl import load_workbook


def check_excel_exist_relation(subjects, fname, sheetname):
    if sheetname == 'Cortical':
        col_names = ['Subject', 'N20', 'N20_amplitude', 'N20_high',
                     'N20_high_amplitude', 'P39', 'P39_amplitude', 'P39_high',
                     'P39_high_amplitude']
    elif sheetname == 'Thalamic':
        col_names = ['Subject', 'P14', 'P14_amplitude', 'P14_high',
                     'P14_high_amplitude', 'P30', 'P30_amplitude', 'P30_high',
                     'P30_high_amplitude']
    elif sheetname == 'Spinal':
        col_names = ['Subject', 'N13', 'N13_amplitude', 'N13_high',
                     'N13_high_amplitude', 'N22', 'N22_amplitude', 'N22_high',
                     'N22_high_amplitude']
    # If the excel workbook doesn't exist, create it
    if not os.path.isfile(fname):
        workbook = xlsxwriter.Workbook(fname)
        worksheet = workbook.add_worksheet(name=sheetname)
        i = 0
        for col in col_names:
            worksheet.write(0, i, col)
            i += 1
        j = 1
        for subject in subjects:
            worksheet.write(j, 0, subject)
            j += 1
        workbook.close()
    # If it does exist, check the worksheet exists, and if not, add it
    else:
        wb = load_workbook(filename=fname)
        if sheetname not in wb.sheetnames:
            worksheet = wb.create_sheet(sheetname)
            i = 1
            for col in col_names:
                worksheet.cell(row=1, column=i).value = col
                i += 1
            j = 2
            for subject in subjects:
                worksheet.cell(row=j, column=1).value = subject
                j += 1
            wb.save(fname)