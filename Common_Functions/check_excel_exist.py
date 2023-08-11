import os
import xlsxwriter
from openpyxl import load_workbook


def check_excel_exist(srmr_nr, subjects, component_fname, component_sheetname, visibility_fname, visibility_sheetname,
                      digits=False):
    # Check component selection sheet
    if srmr_nr == 1:
        col_names = ['Subject', 'sigma_median_comp', 'sigma_median_flip', 'sigma_tibial_comp', 'sigma_tibial_flip']
    else:
        if digits:
            col_names = ['Subject', 'sigma_med_digits_comp', 'sigma_med_digits_flip', 'sigma_tib_digits_comp',
                         'sigma_tib_digits_flip']
        else:
            col_names = ['Subject', 'sigma_med_mixed_comp', 'sigma_med_mixed_flip', 'sigma_tib_mixed_comp',
                         'sigma_tib_mixed_flip']
    # If the excel workbook doesn't exist, create it
    if not os.path.isfile(component_fname):
        workbook = xlsxwriter.Workbook(component_fname)
        worksheet = workbook.add_worksheet(name=component_sheetname)
        i = 0
        for col in col_names:
            worksheet.write(0, i, col)
            i += 1
        j = 1
        for subject in subjects:
            worksheet.write(j, 0, subject)
            j += 1
        workbook.close()
    # If it does exist, check the worksheet exists, and if not, add it
    else:
        wb = load_workbook(filename=component_fname)
        if component_sheetname not in wb.sheetnames:
            worksheet = wb.create_sheet(component_sheetname)
            i = 1
            for col in col_names:
                worksheet.cell(row=1, column=i).value = col
                i += 1
            j = 2
            for subject in subjects:
                worksheet.cell(row=j, column=1).value = subject
                j += 1
            wb.save(component_fname)

    # Check visibility sheet
    if srmr_nr == 1:
        col_names_vis = ['Subject', 'Sigma_Median_Visible', 'Sigma_Tibial_Visible']
    else:
        if digits:
            col_names_vis = ['Subject', 'Sigma_Med_digits_Visible', 'Sigma_Tib_digits_Visible']
        else:
            col_names_vis = ['Subject', 'Sigma_Med_mixed_Visible', 'Sigma_Tib_mixed_Visible']
    if not os.path.isfile(visibility_fname):
        workbook = xlsxwriter.Workbook(visibility_fname)
        worksheet = workbook.add_worksheet(name=visibility_sheetname)
        i = 0
        for col in col_names_vis:
            worksheet.write(0, i, col)
            i += 1
        j = 1
        for subject in subjects:
            worksheet.write(j, 0, subject)
            j += 1
        workbook.close()
    # If it does exist, check the worksheet exists, and if not, add it
    else:
        wb = load_workbook(filename=visibility_fname)
        if component_sheetname not in wb.sheetnames:
            worksheet = wb.create_sheet(visibility_sheetname)
            i = 1
            for col in col_names_vis:
                worksheet.cell(row=1, column=i).value = col
                i += 1
            j = 2
            for subject in subjects:
                worksheet.cell(row=j, column=1).value = subject
                j += 1
            wb.save(visibility_fname)
    return